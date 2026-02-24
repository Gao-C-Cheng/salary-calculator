from operator import index

from openpyxl import load_workbook
import os
from pandas import DataFrame, ExcelWriter
from openpyxl.utils import get_column_letter
import logging


logger = logging.getLogger(__name__)


def copy_and_get_header_openpyxl(source_file, target_file, header_rows=1):
    """
    复制Excel文件，保留表头及其完整格式，清除其余数据

    参数:
        source_file (str): 源Excel文件路径
        target_file (str): 目标Excel文件路径
        header_rows (int): 要保留的表头行数，默认为1
    """
    try:
        # 加载源工作簿
        source_wb = load_workbook(source_file)
        source_ws = source_wb.active  # 获取活动工作表

        # 创建新工作簿
        target_wb = load_workbook(source_file)  # 通过加载源文件来完整复制工作簿结构
        target_ws = target_wb.active

        # 获取最大行数
        max_row = source_ws.max_row

        # 从最后一行开始向上删除，避免行号变化导致的问题
        # 只保留从header_rows+1开始到底部的行
        if max_row > header_rows:
            target_ws.delete_rows(header_rows + 1, max_row - header_rows)

        # 确保目标目录存在
        os.makedirs(os.path.dirname(target_file), exist_ok=True)

        # 保存新文件
        target_wb.save(target_file)
        print(f"文件已处理完成！保留前{header_rows}行表头及其完整格式。输出文件：{target_file}")

    except FileNotFoundError:
        print(f"错误：找不到源文件 '{source_file}'，请检查文件路径。")
    except Exception as e:
        print(f"处理过程中发生错误：{e}")


def append_data_to_excel_openpyxl(file_path, data: DataFrame, start_row):
    """
    使用openpyxl将数据追加到Excel文件的指定起始行

    参数:
        file_path (str): 目标Excel文件路径
        data (list of list): 要追加的数据，每个子列表表示一行
        start_row (int): 数据开始写入的行号（1基准）
    """
    # 加载工作表
    workbook = load_workbook(file_path)
    # 获取第一个sheet名
    sheet_name = workbook.sheetnames[0]
    workbook.close()
    logger.info(f"准备写入excel表：{sheet_name}")
    with ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        data.to_excel(
            writer,
            sheet_name=sheet_name,  # 替换为实际工作表名
            startrow=start_row,  # 从第6行开始（0-based索引）
            header=False,
            index=False
        )
        # 获取工作簿和工作表对象
        workbook = writer.book
        worksheet = workbook[sheet_name]

        # 遍历每一列，自动调整列宽
        for column_cells in worksheet.columns:
            max_length = 0
            try:
                column_letter = get_column_letter(column_cells[0].column)  # 尝试获取列字母
            except AttributeError:
                # 如果col[0]是MergedCell，会触发AttributeError，跳过此列
                print(f"跳过合并单元格所在列，起始单元格坐标: {column_cells[0].coordinate}")
                continue
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception as e:
                    logger.error(f"计算列宽时出错: {e}")
            adjusted_width = (max_length + 2)  # 添加一些额外空间
            worksheet.column_dimensions[column_letter].width = adjusted_width



def export_dataframe_with_proper_column_width(file_path, data: DataFrame, header=True):
    """
    导出DataFrame到Excel，并自动调整列宽以适应内容

    参数:
        file_path (str): 目标Excel文件路径
        data (DataFrame): 要导出的DataFrame
    """
    # 确保目标目录存在
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    with ExcelWriter(file_path, engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Sheet1', header=header)

        # 获取工作簿和工作表对象
        workbook = writer.book
        worksheet = workbook['Sheet1']

        # 遍历每一列，自动调整列宽
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter  # 获取列字母
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception as e:
                    logger.error(f"计算列宽时出错: {e}")
            adjusted_width = (max_length*1.5 + 2)  # 添加一些额外空间
            worksheet.column_dimensions[column_letter].width = adjusted_width

    logger.info(f"数据已成功导出到Excel文件：{file_path}，并调整了列宽。")